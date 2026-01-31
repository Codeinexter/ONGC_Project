import os
import random
import sqlite3
import bcrypt

from flask import Flask, render_template, request, redirect, url_for, session, flash
from flask_mail import Mail, Message
from werkzeug.utils import secure_filename
# from twilio.rest import Client
from dotenv import load_dotenv
from PIL import Image
from flask import abort
from flask import send_from_directory
from datetime import datetime
from zoneinfo import ZoneInfo
import pytz
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from openpyxl import Workbook
from reportlab.lib.units import inch

# Load Environment Variables
load_dotenv()

# Flask App Setup
app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY")

app.config["MAX_CONTENT_LENGTH"] = 2 * 1024 * 1024  # 2 MB limit

@app.errorhandler(413)
def file_too_large(e):
    flash("Image too large! Max size allowed is 2MB.", "danger")
    return redirect(request.url)

@app.errorhandler(403)
def forbidden(e):
    flash("You do not have permission to access this file.", "danger")
    return redirect(url_for("dashboard"))

# Timezone Conversion
def utc_to_ist(timestamp):
    # timestamp from SQLite: 'YYYY-MM-DD HH:MM:SS'
    dt_utc = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
    dt_utc = dt_utc.replace(tzinfo=ZoneInfo("UTC"))
    dt_ist = dt_utc.astimezone(ZoneInfo("Asia/Kolkata"))

    return dt_ist.strftime("%d %b %Y, %I:%M %p")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE_PATH = os.path.join(BASE_DIR, "database.db")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "images")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

DATA_UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "uploads")
os.makedirs(DATA_UPLOAD_FOLDER, exist_ok=True)
app.config["DATA_UPLOAD_FOLDER"] = DATA_UPLOAD_FOLDER

ALLOWED_EXTENSIONS = {
    "png", "jpg", "jpeg", "gif",
    "pdf", "docx", "xlsx", "csv",
    "zip", "rar", "txt"
}

def allowed_file(filename):
    return "." in filename and \
            filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

# Flask-Mail Configuration
app.config.update(
    MAIL_SERVER=os.getenv("MAIL_SERVER"),
    MAIL_PORT=int(os.getenv("MAIL_PORT")),
    MAIL_USE_TLS=os.getenv("MAIL_USE_TLS") == "True",
    MAIL_USERNAME=os.getenv("MAIL_USERNAME"),
    MAIL_PASSWORD=os.getenv("MAIL_PASSWORD"),
    MAIL_DEFAULT_SENDER=os.getenv("MAIL_USERNAME")
)

mail = Mail(app)

# Twilio Configuration
# twilio_client = Client(
#     os.getenv("TWILIO_ACCOUNT_SID"),
#     os.getenv("TWILIO_AUTH_TOKEN")
# )
# TWILIO_PHONE_NUMBER = os.getenv("TWILIO_PHONE_NUMBER")

# Database Helper
def get_db_connection():
    conn = sqlite3.connect(DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# Utility Functions
def generate_otp():
    return random.randint(100000, 999999)

def send_email_otp(email, otp):
    msg = Message(
        subject="OTP Verification",
        recipients=[email]
    )
    msg.body = f"Your OTP is {otp}. Do not share it."
    mail.send(msg)

# def send_sms_otp(phone, otp):
#     try:
#         twilio_client.messages.create(
#             body=f"Your OTP is {otp}. Do not share it.",
#             from_=TWILIO_PHONE_NUMBER,
#             to=phone
#         )
#     except Exception as e:
#         print("Twilio SMS Error:", e)

# Role hierarchy (higher number = higher privilege)
ROLE_PRIORITY = {
    "user": 1,
    "moderator": 2,
    "admin": 3
}

def can_view_data(viewer_privilege, uploader_privilege, visibility):
    # Admin sees everything
    if viewer_privilege == "admin":
        return True

    # Same privilege always allowed
    if viewer_privilege == uploader_privilege:
        return True

    # Different privilege → check visibility flag
    return ROLE_PRIORITY[viewer_privilege] >= ROLE_PRIORITY[visibility]

def get_next_approver_role(current_role):
    roles = sorted(ROLE_PRIORITY.items(), key=lambda x: x[1])
    for role, priority in roles:
        if priority > ROLE_PRIORITY[current_role]:
            return role
    return None  # No higher role (admin)

# Routes
@app.route("/")
def index():
    return render_template("welcome.html")

# Register
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        name = request.form["name"]
        email = request.form["email"]
        mobile = request.form["mobile_number"]
        password = request.form["password"]
        privilege = request.form["privilege"]
        image = request.files.get("image")

        conn = get_db_connection()

        # check for duplicate email
        existing_user = conn.execute(
            "SELECT id FROM user WHERE email = ?",
            (email,)
        ).fetchone()

        if existing_user:
            conn.close()
            flash("Email already registered. Please login.", "danger")
            return redirect(url_for("register"))

        # check duplicate mobile
        existing_mobile = conn.execute(
            "SELECT id FROM user WHERE mobile_number = ?",
            (mobile,)
        ).fetchone()

        if existing_mobile:
            conn.close()
            flash("Mobile number already registered.", "danger")
            return redirect(url_for("register"))

        conn.close()

        hashed_password = bcrypt.hashpw(
            password.encode("utf-8"),
            bcrypt.gensalt()
        )

        filename = "default.png"
        if image and image.filename:
            filename = secure_filename(image.filename)
            image_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)

            img = Image.open(image)
            img = img.convert("RGB")                 # Safety
            img.thumbnail((300, 300))                # Max size 300x300
            img.save(image_path, optimize=True, quality=85)

        email_otp = generate_otp()
#       mobile_otp = generate_otp()

        session["email_otp"] = email_otp
#       session["mobile_otp"] = mobile_otp
        session["user_data"] = {
            "name": name,
            "email": email,
            "mobile": mobile,
            "password": hashed_password,
            "privilege": privilege,
            "image": filename
        }

        send_email_otp(email, email_otp)
#       send_sms_otp(mobile, mobile_otp)

        return redirect(url_for("verify"))

    return render_template("register.html")

# Verify OTP
@app.route("/verify", methods=["GET", "POST"])
def verify():
    if request.method == "POST":
        email_otp_input = int(request.form["email_otp"])
        #mobile_otp_input = int(request.form["mobile_otp"])

        if email_otp_input != session.get("email_otp"):
            flash("Invalid OTP", "danger")
            return redirect(url_for("verify"))
        
        # if mobile_otp_input != session.get("mobile_otp"):
        #     flash("Invalid OTP", "danger")
        #     return redirect(url_for("verify"))

        session["email_verified"] = True
        #session["mobile_verified"] = True
        
        user = session["user_data"]

        conn = get_db_connection()
        conn.execute("""
            INSERT INTO user (name, email, mobile_number, password, privilege, email_verified, mobile_verified, image_path)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            user["name"],
            user["email"],
            user["mobile"],
            user["password"],
            user["privilege"],
            1,
            1,
            user["image"]
        ))
        conn.commit()
        conn.close()

        session.clear()
        flash("Account verified successfully!", "success")
        return redirect(url_for("login"))

    return render_template("verify.html")

# Login
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form["email"]
        password = request.form["password"]

        conn = get_db_connection()
        user = conn.execute(
            "SELECT * FROM user WHERE email = ? AND email_verified = 1 AND mobile_verified = 1",
            (email,)
        ).fetchone()
        conn.close()

        if user and bcrypt.checkpw(password.encode("utf-8"), user["password"]):
            session["user_id"] = user["id"]
            return redirect(url_for("dashboard"))

        flash("Invalid credentials or unverified account", "danger")

    return render_template("login.html")

# Home
@app.route("/home")
def home():
    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()
    user = conn.execute(
        "SELECT name, privilege, image_path FROM user WHERE id = ?",
        (session["user_id"],)
    ).fetchone()
    conn.close()

    return render_template(
        "home.html",
        name=user["name"],
        privilege=user["privilege"],
        image_url=f"/static/images/{user['image_path']}"
    )

# Dashboard
@app.route("/dashboard")
def dashboard():
    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()

    current_user = conn.execute(
        "SELECT id, privilege FROM user WHERE id = ?",
        (session["user_id"],)
    ).fetchone()
    
    if not current_user:
        conn.close()
        session.clear()
        abort(403)

    role = current_user["privilege"]
    user_id = current_user["id"]

    # Approved files (existing logic)
    approved_files = conn.execute("""
        SELECT d.*, u.privilege AS uploader_privilege
        FROM data_items d
        JOIN user u ON d.uploaded_by = u.id
        WHERE d.approval_status = 'approved'
        ORDER BY d.created_at DESC
    """).fetchall()

    # User: pending uploads
    pending_user_files = []
    rejected_files = []
    if role == "user":
        pending_user_files = conn.execute("""
            SELECT *
            FROM data_items
            WHERE uploaded_by = ?
                AND approval_status = 'pending'
            ORDER BY created_at DESC
        """, (user_id,)).fetchall()

        rejected_files = conn.execute("""
            SELECT *
            FROM data_items
            WHERE uploaded_by = ?
                AND approval_status = 'rejected'
            ORDER BY created_at DESC
        """, (user_id,)).fetchall()

    # Moderator: approvals + waiting for admin
    moderator_pending = []
    moderator_waiting_admin = []

    if role == "moderator":
        moderator_pending = conn.execute("""
            SELECT ar.*, d.filename, u.name
            FROM approval_requests ar
            JOIN data_items d ON ar.data_id = d.id
            JOIN user u ON ar.requested_by = u.id
            WHERE ar.current_approver_role = 'moderator'
                AND ar.status = 'pending'
        """).fetchall()

        moderator_waiting_admin = conn.execute("""
            SELECT d.*
            FROM data_items d
            WHERE d.uploaded_by = ?
                AND d.approval_status = 'pending'
        """, (user_id,)).fetchall()

    # Admin: approvals
    admin_pending = []
    if role == "admin":
        admin_pending = conn.execute("""
            SELECT ar.*, d.filename, u.name
            FROM approval_requests ar
            JOIN data_items d ON ar.data_id = d.id
            JOIN user u ON ar.requested_by = u.id
            WHERE ar.current_approver_role = 'admin'
                AND ar.status = 'pending'
        """).fetchall()

    all_data = conn.execute("""
        SELECT d.*, u.privilege AS uploader_privilege
        FROM data_items d
        JOIN user u ON d.uploaded_by = u.id
        WHERE d.approval_status = 'approved'
        ORDER BY d.created_at DESC
    """).fetchall()

    conn.close()

    visible_data = []

    for item in all_data:
        if can_view_data(
            current_user["privilege"],
            item["uploader_privilege"],
            item["visibility"]
        ):
            item = dict(item)
            item["created_at"] = utc_to_ist(item["created_at"])
            visible_data.append(item)

    return render_template(
        "dashboard.html",
        data_items=visible_data,
        privilege=current_user["privilege"],
        role=role,
        approved_files=approved_files,
        pending_user_files=pending_user_files,
        rejected_files=rejected_files,
        moderator_pending=moderator_pending,
        moderator_waiting_admin=moderator_waiting_admin,
        admin_pending=admin_pending
    )

# Upload Data
from uuid import uuid4
@app.route("/upload", methods=["GET", "POST"])
def upload():
    if "user_id" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":
        file = request.files.get("file")
        visibility = request.form["visibility"]

        if not file or file.filename == "":
            flash("No file selected", "danger")
            return redirect(request.url)

        if not allowed_file(file.filename):
            flash("File type not allowed", "danger")
            return redirect(request.url)

        original_name = secure_filename(file.filename)
        unique_name = f"{uuid4().hex}_{original_name}"

        file_path = os.path.join(
            app.config["DATA_UPLOAD_FOLDER"],
            unique_name
        )
        file.save(file_path)

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            """
            INSERT INTO data_items
            (filename, stored_filename, file_type, uploaded_by, visibility, approval_status)
            VALUES (?, ?, ?, ?, ?, 'pending')
            """,
            (
                original_name,
                unique_name,
                original_name.rsplit(".", 1)[1].lower(),
                session["user_id"],
                visibility
            )
        )

        # Capture inserted data_items ID
        data_id = cursor.lastrowid

        if not data_id:
            conn.rollback()
            conn.close()
            raise Exception("data_id was not generated")

        # get uploader role
        uploader = conn.execute(
            "SELECT privilege FROM user WHERE id = ?",
            (session["user_id"],)
        ).fetchone()

        next_role = get_next_approver_role(uploader["privilege"])

        if next_role:
            conn.execute("""
                INSERT INTO approval_requests
                (data_id, requested_by, current_approver_role)
                VALUES (?, ?, ?)
            """, (
                    data_id,
                    session["user_id"],
                    next_role
                )
            )
        else:
        # Admin upload → auto approved
            conn.execute("""
            UPDATE data_items
            SET approval_status = 'approved'
            WHERE id = ?
            """, (
                    data_id,
                )
            )

        conn.commit()
        conn.close()

        flash("File uploaded successfully", "success")
        return redirect(url_for("dashboard"))

    return render_template("upload.html")

# Approval Requests
@app.route("/approvals")
def approvals():
    if "user_id" not in session:
        return redirect(url_for("login"))

    conn = get_db_connection()

    role = conn.execute(
        "SELECT privilege FROM user WHERE id = ?",
        (session["user_id"],)
    ).fetchone()["privilege"]

    requests = conn.execute("""
        SELECT ar.*, d.filename, d.stored_filename, u.name
        FROM approval_requests ar
        JOIN data_items d ON ar.data_id = d.id
        JOIN user u ON ar.requested_by = u.id
        WHERE ar.current_approver_role = ?
            AND ar.status = 'pending'
    """, (role,)).fetchall()

    conn.close()
    return render_template("approvals.html", requests=requests)

# Approve Request
@app.route("/approve/<int:req_id>", methods=["POST"])
def approve(req_id):
    if "user_id" not in session:
        abort(403)

    conn = get_db_connection()
    cursor = conn.cursor()

    # Get approver role
    approver = conn.execute(
        "SELECT privilege FROM user WHERE id = ?",
        (session["user_id"],)
    ).fetchone()

    if not approver:
        conn.close()
        abort(403)

    # Get approval request
    req = conn.execute("""
        SELECT ar.*, d.uploaded_by
        FROM approval_requests ar
        JOIN data_items d ON ar.data_id = d.id
        WHERE ar.id = ?
            AND ar.status = 'pending'
    """, (req_id,)).fetchone()

    if not approver or not req:
        conn.close()
        abort(404)

    # MODERATOR approval → forward to admin
    if approver["privilege"] == "moderator":
        # Forward to admin
        cursor.execute("""
            UPDATE approval_requests
            SET current_approver_role = 'admin'
            WHERE id = ?
        """, (req_id,))

    # ADMIN approval → final approval
    elif approver["privilege"] == "admin":
        # Mark approval request complete
        cursor.execute("""
            UPDATE approval_requests
            SET status = 'approved'
            WHERE id = ?
        """, (req_id,))

        cursor.execute("""
            UPDATE data_items
            SET approval_status = 'approved',
                visibility = 'user'
            WHERE id = ?
        """, (req["data_id"],))

    conn.commit()
    conn.close()
    flash("Request approved", "success")
    return redirect(url_for("approvals"))

# Reject Request
@app.route("/reject/<int:req_id>", methods=["POST"])
def reject(req_id):
    if "user_id" not in session:
        abort(403)

    remark = request.form["remark"].strip()

    conn = get_db_connection()
    cursor = conn.cursor()

    approver = conn.execute(
        "SELECT privilege FROM user WHERE id = ?",
        (session["user_id"],)
    ).fetchone()

    req = conn.execute("""
        SELECT ar.id, ar.data_id
        FROM approval_requests ar
        WHERE ar.id = ?
            AND ar.status = 'pending'
    """, (req_id,)).fetchone()

    if not approver or not req:
        conn.close()
        abort(404)

    conn.execute("""
        UPDATE approval_requests
        SET status = 'rejected',
            remarks = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (remark, req_id))

    # Update file status
    cursor.execute("""
        UPDATE data_items
        SET approval_status = 'rejected'
        WHERE id = ?
    """, (req["data_id"],))

    conn.commit()
    conn.close()
    flash("Request rejected", "danger")
    return redirect(url_for("approvals"))

# Preview File
from flask import send_from_directory, abort
@app.route("/file/<path:stored_filename>")
def serve_file(stored_filename):
    if "user_id" not in session:
        abort(403)
    
    conn = get_db_connection()

    item = conn.execute("""
        SELECT d.*, u.privilege AS uploader_privilege
        FROM data_items d
        JOIN user u ON d.uploaded_by = u.id
        WHERE d.stored_filename = ?
            AND d.approval_status = 'approved'
    """, (stored_filename,)).fetchone()

    current_user = conn.execute(
        "SELECT privilege FROM user WHERE id = ?",
        (session["user_id"],)
    ).fetchone()

    conn.close()

    if not item or not current_user:
        abort(404)

    # Approved files → normal visibility rules
    if item["approval_status"] == "approved":
        if not can_view_data(
            current_user["privilege"],
            item["uploader_privilege"],
            item["visibility"]
        ):
            abort(403)

    # Pending / Rejected files → ONLY moderator/admin can preview
    else:
        if current_user["privilege"] not in ["moderator", "admin"]:
            abort(403)

    file_path = os.path.join(
        app.config["DATA_UPLOAD_FOLDER"],
        stored_filename
    )

    if not os.path.exists(file_path):
        abort(404)

    return send_from_directory(
        app.config["DATA_UPLOAD_FOLDER"],
        stored_filename,
        as_attachment=False
    )

# Delete File
@app.route("/delete-file/<int:file_id>", methods=["POST"])
def delete_file(file_id):
    if "user_id" not in session:
        abort(403)

    conn = get_db_connection()

    # Get current user privilege
    current_user = conn.execute(
        "SELECT id, privilege FROM user WHERE id = ?",
        (session["user_id"],)
    ).fetchone()

    # Get file info
    file = conn.execute(
        "SELECT * FROM data_items WHERE id = ?",
        (file_id,)
    ).fetchone()

    if not file:
        conn.close()
        abort(404)

    # Authorization rule:
    # uploader OR admin
    if (
        file["uploaded_by"] != current_user["id"]
        and current_user["privilege"] != "admin"
    ):
        conn.close()
        abort(403)

    # Delete physical file
    file_path = os.path.join(
        app.config["DATA_UPLOAD_FOLDER"],
        file["stored_filename"]
    )
    if os.path.exists(file_path):
        os.remove(file_path)

    # Delete DB record
    conn.execute(
        "DELETE FROM data_items WHERE id = ?",
        (file_id,)
    )
    conn.commit()
    conn.close()

    flash("File deleted successfully", "success")
    return redirect(url_for("dashboard"))

@app.route("/delete-pending/<int:data_id>", methods=["POST"])
def delete_pending(data_id):
    if "user_id" not in session:
        abort(403)

    conn = get_db_connection()
    cursor = conn.cursor()

    # Get current user
    user = conn.execute(
        "SELECT id, privilege FROM user WHERE id = ?",
        (session["user_id"],)
    ).fetchone()

    # Get file info
    file = conn.execute("""
        SELECT *
        FROM data_items
        WHERE id = ?
    """, (data_id,)).fetchone()

    if not file:
        conn.close()
        abort(404)

    # Only pending files can be deleted
    if file["approval_status"] not in ["pending", "rejected"]:
        conn.close()
        abort(403)

    # Only uploader can delete
    if file["uploaded_by"] != user["id"]:
        conn.close()
        abort(403)

    # Delete physical file
    file_path = os.path.join(
        app.config["DATA_UPLOAD_FOLDER"],
        file["stored_filename"]
    )
    if os.path.exists(file_path):
        os.remove(file_path)

    # Delete approval request
    cursor.execute(
        "DELETE FROM approval_requests WHERE data_id = ?",
        (data_id,)
    )

    # Delete data record
    cursor.execute(
        "DELETE FROM data_items WHERE id = ?",
        (data_id,)
    )

    conn.commit()
    conn.close()

    flash("File deleted successfully", "success")
    return redirect(url_for("dashboard"))

# Logout
@app.route("/logout")
def logout():
    session.clear()
    flash("Logged out successfully", "info")
    return redirect(url_for("login"))

# Reports Data
def get_report_data():
    conn = get_db_connection()

    total_users = conn.execute(
        "SELECT COUNT(*) FROM user"
    ).fetchone()[0]

    users_by_privilege = conn.execute("""
        SELECT privilege, COUNT(*) AS count
        FROM user
        GROUP BY privilege
    """).fetchall()

    uploads_by_type = conn.execute("""
        SELECT COALESCE(file_type, 'unknown') AS file_type,
                COUNT(*) AS count
        FROM data_items
        GROUP BY file_type
    """).fetchall()

    monthly_uploads = conn.execute("""
        SELECT substr(created_at, 1, 7) AS month,
                COUNT(*) AS count
        FROM data_items
        GROUP BY month
        ORDER BY month
    """).fetchall()

    monthly_requests = conn.execute("""
        SELECT substr(created_at, 1, 7) AS month,
                COUNT(*) AS count
        FROM approval_requests
        GROUP BY month
        ORDER BY month
    """).fetchall()

    requests_by_status = conn.execute("""
        SELECT status, COUNT(*) AS count
        FROM approval_requests
        GROUP BY status
    """).fetchall()

    requests_by_role = conn.execute("""
        SELECT current_approver_role, COUNT(*) AS count
        FROM approval_requests
        GROUP BY current_approver_role
    """).fetchall()

    conn.close()

    return {
        "total_users": total_users,
        "users_by_privilege": users_by_privilege,
        "uploads_by_type": uploads_by_type,
        "monthly_uploads": monthly_uploads,
        "monthly_requests": monthly_requests,
        "requests_by_status": requests_by_status,
        "requests_by_role": requests_by_role
    }

# Generate PDF Report
@app.route("/report/pdf")
def generate_pdf_report():
    if "user_id" not in session:
        return redirect(url_for("login"))

    data = get_report_data()

    filename = "ongc_system_report.pdf"
    filepath = os.path.join(BASE_DIR, filename)

    doc = SimpleDocTemplate(
        filepath,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("<b>ONGC Data & Approval System Report</b>", styles["Title"]))
    elements.append(Spacer(1, 0.3 * inch))

    # Summary
    elements.append(Paragraph("<b>System Summary</b>", styles["Heading2"]))
    elements.append(Paragraph(f"Total Registered Users: {data['total_users']}", styles["Normal"]))
    elements.append(Spacer(1, 0.2 * inch))

    # Users by Privilege
    elements.append(Paragraph("<b>Users by Privilege</b>", styles["Heading2"]))
    table_data = [["Privilege", "Count"]]
    for row in data["users_by_privilege"]:
        table_data.append([row["privilege"], row["count"]])
    elements.append(Table(table_data))
    elements.append(Spacer(1, 0.2 * inch))

    # Monthly Uploads
    elements.append(Paragraph("<b>Monthly File Uploads</b>", styles["Heading2"]))
    table_data = [["Month", "Uploads"]]
    for row in data["monthly_uploads"]:
        table_data.append([row["month"], row["count"]])
    elements.append(Table(table_data))
    elements.append(Spacer(1, 0.2 * inch))

    # Uploads by File Type
    elements.append(Paragraph("<b>Uploads by File Type</b>", styles["Heading2"]))
    table_data = [["File Type", "Count"]]
    for row in data["uploads_by_type"]:
        table_data.append([row["file_type"], row["count"]])
    elements.append(Table(table_data))
    elements.append(Spacer(1, 0.2 * inch))

    # Monthly Approval Requests
    elements.append(Paragraph("<b>Monthly Approval Requests</b>", styles["Heading2"]))
    table_data = [["Month", "Requests"]]
    for row in data["monthly_requests"]:
        table_data.append([row["month"], row["count"]])
    elements.append(Table(table_data))
    elements.append(Spacer(1, 0.2 * inch))

    # Requests by Status
    elements.append(Paragraph("<b>Approval Requests by Status</b>", styles["Heading2"]))
    table_data = [["Status", "Count"]]
    for row in data["requests_by_status"]:
        table_data.append([row["status"], row["count"]])
    elements.append(Table(table_data))
    elements.append(Spacer(1, 0.2 * inch))

    # Requests by Approver Role
    elements.append(Paragraph("<b>Requests by Approver Role</b>", styles["Heading2"]))
    table_data = [["Approver Role", "Count"]]
    for row in data["requests_by_role"]:
        table_data.append([row["current_approver_role"], row["count"]])
    elements.append(Table(table_data))

    doc.build(elements)

    return send_from_directory(BASE_DIR, filename, as_attachment=True)

# Generate Excel Report
@app.route("/report/excel")
def generate_excel_report():
    if "user_id" not in session:
        return redirect(url_for("login"))

    data = get_report_data()
    wb = Workbook()

    # Summary Sheet
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Metric", "Value"])
    ws.append(["Total Users", data["total_users"]])

    # Users by Privilege
    ws_role = wb.create_sheet("Users by Privilege")
    ws_role.append(["Privilege", "Count"])
    for row in data["users_by_privilege"]:
        ws_role.append([row["privilege"], row["count"]])

    # Monthly Uploads
    ws_uploads = wb.create_sheet("Monthly Uploads")
    ws_uploads.append(["Month", "Uploads"])
    for row in data["monthly_uploads"]:
        ws_uploads.append([row["month"], row["count"]])

    # Uploads by File Type
    ws_types = wb.create_sheet("Uploads by File Type")
    ws_types.append(["File Type", "Count"])
    for row in data["uploads_by_type"]:
        ws_types.append([row["file_type"], row["count"]])

    # Monthly Requests
    ws_requests = wb.create_sheet("Monthly Requests")
    ws_requests.append(["Month", "Requests"])
    for row in data["monthly_requests"]:
        ws_requests.append([row["month"], row["count"]])

    # Requests by Status
    ws_status = wb.create_sheet("Requests by Status")
    ws_status.append(["Status", "Count"])
    for row in data["requests_by_status"]:
        ws_status.append([row["status"], row["count"]])

    # Requests by Role
    ws_role_req = wb.create_sheet("Requests by Approver Role")
    ws_role_req.append(["Approver Role", "Count"])
    for row in data["requests_by_role"]:
        ws_role_req.append([row["current_approver_role"], row["count"]])

    filename = "ongc_system_report.xlsx"
    filepath = os.path.join(BASE_DIR, filename)
    wb.save(filepath)

    return send_from_directory(BASE_DIR, filename, as_attachment=True)

# Run App
if __name__ == "__main__":
    app.run(debug=True)